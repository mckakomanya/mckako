#!/usr/bin/env python3
"""
Generador de Plan de Ahorro e Inversión
Para pareja de médicos residentes en Uruguay.
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, BarChart, Reference
from copy import copy

# ─── Color palette ───────────────────────────────────────────────────────────
WHITE = "FFFFFF"
DARK_NAVY = "1B2A4A"
MEDIUM_BLUE = "2E5090"
LIGHT_BLUE = "4472C4"
SOFT_BLUE = "D6E4F0"
ACCENT_GREEN = "70AD47"
ACCENT_ORANGE = "ED7D31"
ACCENT_RED = "C00000"
ACCENT_YELLOW = "FFC000"
LIGHT_GRAY = "F2F2F2"
MEDIUM_GRAY = "D9D9D9"
DARK_GRAY = "595959"

# ─── Reusable styles ────────────────────────────────────────────────────────
thin_border = Border(
    left=Side(style="thin", color=MEDIUM_GRAY),
    right=Side(style="thin", color=MEDIUM_GRAY),
    top=Side(style="thin", color=MEDIUM_GRAY),
    bottom=Side(style="thin", color=MEDIUM_GRAY),
)

header_font = Font(name="Calibri", bold=True, size=11, color=WHITE)
header_fill = PatternFill(start_color=MEDIUM_BLUE, end_color=MEDIUM_BLUE, fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

title_font = Font(name="Calibri", bold=True, size=16, color=DARK_NAVY)
subtitle_font = Font(name="Calibri", bold=True, size=13, color=MEDIUM_BLUE)
section_font = Font(name="Calibri", bold=True, size=11, color=MEDIUM_BLUE)

currency_format = '#,##0'
pct_format = '0.0%'

MESES = [
    "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
]


def style_header_row(ws, row, col_start, col_end):
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border


def style_data_cell(ws, row, col, fmt=None):
    cell = ws.cell(row=row, column=col)
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if fmt:
        cell.number_format = fmt
    return cell


def style_range(ws, row_start, row_end, col_start, col_end, fmt=None):
    for r in range(row_start, row_end + 1):
        for c in range(col_start, col_end + 1):
            style_data_cell(ws, r, c, fmt)


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_section_title(ws, row, col, text):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = subtitle_font
    cell.alignment = Alignment(vertical="center")


def write_label(ws, row, col, text, bold=False):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(name="Calibri", bold=bold, size=10, color=DARK_GRAY)
    cell.border = thin_border
    cell.alignment = Alignment(vertical="center")
    return cell


def alt_row_fill(ws, row, col_start, col_end):
    """Light gray fill for even rows."""
    if row % 2 == 0:
        fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
        for c in range(col_start, col_end + 1):
            ws.cell(row=row, column=c).fill = fill


# ═════════════════════════════════════════════════════════════════════════════
#  SHEET 1: DASHBOARD / RESUMEN
# ═════════════════════════════════════════════════════════════════════════════
def create_dashboard(wb):
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_properties.tabColor = DARK_NAVY
    set_col_widths(ws, [3, 30, 20, 20, 20, 20, 3])

    # Title
    ws.merge_cells("B2:F2")
    cell = ws.cell(row=2, column=2, value="PLAN FINANCIERO - PAREJA MÉDICOS RESIDENTES")
    cell.font = title_font
    cell.alignment = Alignment(horizontal="center")

    ws.merge_cells("B3:F3")
    cell = ws.cell(row=3, column=2, value="Uruguay · Pediatría & Oncología Radioterápica")
    cell.font = Font(name="Calibri", size=11, italic=True, color=DARK_GRAY)
    cell.alignment = Alignment(horizontal="center")

    # ── Section: Resumen de Ingresos ──
    row = 5
    write_section_title(ws, row, 2, "Resumen Mensual")
    row = 6
    labels = ["Ingreso Neto - Pediatría", "Ingreso Neto - Oncología RT",
              "Total Ingresos Pareja", "Total Gastos Mensuales",
              "Disponible para Ahorro", "Tasa de Ahorro"]
    for i, label in enumerate(labels):
        r = row + i
        write_label(ws, r, 2, label, bold=(i >= 2))
        c = ws.cell(row=r, column=3)
        c.border = thin_border
        c.alignment = Alignment(horizontal="center")
        if i == 0:
            c.value = "=Ingresos!D10"
            c.number_format = currency_format
        elif i == 1:
            c.value = "=Ingresos!E10"
            c.number_format = currency_format
        elif i == 2:
            c.value = "=Ingresos!F10"
            c.number_format = currency_format
            c.font = Font(bold=True)
        elif i == 3:
            c.value = "=Gastos!H43"
            c.number_format = currency_format
        elif i == 4:
            c.value = f"=C{row+2}-C{row+3}"
            c.number_format = currency_format
            c.font = Font(bold=True, color=ACCENT_GREEN)
        elif i == 5:
            c.value = f'=IF(C{row+2}>0,C{row+4}/C{row+2},0)'
            c.number_format = pct_format

    # ── Section: Objetivos de Ahorro resumen ──
    row = 14
    write_section_title(ws, row, 2, "Progreso de Objetivos")
    row = 15
    headers = ["Objetivo", "Meta ($)", "Ahorrado ($)", "Progreso", "Plazo"]
    for i, h in enumerate(headers):
        ws.cell(row=row, column=2 + i, value=h)
    style_header_row(ws, row, 2, 6)

    obj_names = [
        ("Vacaciones Juntos", "corto"),
        ("Ahorro Individual - Ella", "corto-largo"),
        ("Ahorro Individual - Él", "corto-largo"),
        ("Compra Inmueble", "mediano"),
        ("Retiro / Inversiones", "largo"),
    ]
    for i, (name, plazo) in enumerate(obj_names):
        r = row + 1 + i
        write_label(ws, r, 2, name)
        # Meta
        c = ws.cell(row=r, column=3)
        c.value = f"=Objetivos!D{6+i}"
        c.number_format = currency_format
        c.border = thin_border
        c.alignment = Alignment(horizontal="center")
        # Ahorrado
        c = ws.cell(row=r, column=4)
        c.value = f"=Objetivos!E{6+i}"
        c.number_format = currency_format
        c.border = thin_border
        c.alignment = Alignment(horizontal="center")
        # Progreso
        c = ws.cell(row=r, column=5)
        c.value = f'=IF(C{r}>0,D{r}/C{r},0)'
        c.number_format = pct_format
        c.border = thin_border
        c.alignment = Alignment(horizontal="center")
        # Plazo
        write_label(ws, r, 6, plazo)
        alt_row_fill(ws, r, 2, 6)

    # ── Section: Regla de distribución ──
    row = 23
    write_section_title(ws, row, 2, "Distribución Sugerida del Ahorro")
    row = 24
    dist_items = [
        ("Fondo de Emergencia (3-6 meses gastos)", "20%"),
        ("Vacaciones", "10%"),
        ("Ahorro Individual (c/u)", "15%"),
        ("Inmueble", "30%"),
        ("Retiro / Inversiones", "25%"),
    ]
    headers2 = ["Categoría", "% Sugerido", "Monto Mensual"]
    for i, h in enumerate(headers2):
        ws.cell(row=row, column=2 + i, value=h)
    style_header_row(ws, row, 2, 4)

    for i, (cat, pct) in enumerate(dist_items):
        r = row + 1 + i
        write_label(ws, r, 2, cat)
        c = ws.cell(row=r, column=3, value=float(pct.strip('%')) / 100)
        c.number_format = pct_format
        c.border = thin_border
        c.alignment = Alignment(horizontal="center")
        c = ws.cell(row=r, column=4)
        c.value = f"=C{r}*Dashboard!C10"
        c.number_format = currency_format
        c.border = thin_border
        c.alignment = Alignment(horizontal="center")
        alt_row_fill(ws, r, 2, 4)

    # ── Notas ──
    row = 32
    write_section_title(ws, row, 2, "Notas")
    notes = [
        "• Los montos están en pesos uruguayos (UYU). Ajustar si usan USD.",
        "• Completar primero las hojas Ingresos y Gastos.",
        "• Los objetivos se configuran en la hoja Objetivos.",
        "• El seguimiento mensual se hace en la hoja Seguimiento.",
        "• Adaptar los % de distribución a su realidad.",
    ]
    for i, note in enumerate(notes):
        c = ws.cell(row=row + 1 + i, column=2, value=note)
        c.font = Font(name="Calibri", size=10, color=DARK_GRAY)

    ws.sheet_view.showGridLines = False


# ═════════════════════════════════════════════════════════════════════════════
#  SHEET 2: INGRESOS
# ═════════════════════════════════════════════════════════════════════════════
def create_ingresos(wb):
    ws = wb.create_sheet("Ingresos")
    ws.sheet_properties.tabColor = ACCENT_GREEN
    set_col_widths(ws, [3, 35, 5, 22, 22, 22])

    ws.merge_cells("B2:F2")
    cell = ws.cell(row=2, column=2, value="INGRESOS MENSUALES")
    cell.font = title_font
    cell.alignment = Alignment(horizontal="center")

    # Headers
    row = 4
    headers = ["Concepto", "", "Pediatría (Ella)", "Oncología RT (Él)", "Total"]
    for i, h in enumerate(headers):
        ws.cell(row=row, column=2 + i, value=h)
    style_header_row(ws, row, 2, 6)

    # Income rows
    items = [
        "Salario Base Residencia",
        "Guardias / Horas Extra",
        "Otros Ingresos (docencia, etc.)",
        "Aguinaldo (prorrateo mensual)",
        "Ingresos Extras",
    ]
    for i, item in enumerate(items):
        r = row + 1 + i
        write_label(ws, r, 2, item)
        style_data_cell(ws, r, 3)  # spacer
        style_data_cell(ws, r, 4, currency_format)
        style_data_cell(ws, r, 5, currency_format)
        c = ws.cell(row=r, column=6)
        c.value = f"=D{r}+E{r}"
        c.number_format = currency_format
        c.border = thin_border
        c.alignment = Alignment(horizontal="center")
        alt_row_fill(ws, r, 2, 6)

    # Total Ingresos Brutos
    r_total_bruto = row + 1 + len(items)
    ws.cell(row=r_total_bruto, column=2, value="TOTAL INGRESOS BRUTOS").font = Font(bold=True, size=11, color=DARK_NAVY)
    for col in [4, 5]:
        c = ws.cell(row=r_total_bruto, column=col)
        c.value = f"=SUM({get_column_letter(col)}{row+1}:{get_column_letter(col)}{r_total_bruto-1})"
        c.number_format = currency_format
        c.font = Font(bold=True)
        c.border = thin_border
        c.alignment = Alignment(horizontal="center")
    c = ws.cell(row=r_total_bruto, column=6)
    c.value = f"=D{r_total_bruto}+E{r_total_bruto}"
    c.number_format = currency_format
    c.font = Font(bold=True)
    c.border = thin_border
    c.alignment = Alignment(horizontal="center")
    style_header_row(ws, r_total_bruto, 2, 6)

    # Deducciones
    r_ded_title = r_total_bruto + 2
    write_section_title(ws, r_ded_title, 2, "Deducciones")
    r_ded_h = r_ded_title + 1
    for i, h in enumerate(headers):
        ws.cell(row=r_ded_h, column=2 + i, value=h)
    style_header_row(ws, r_ded_h, 2, 6)

    deducciones = [
        "BPS (Aporte jubilatorio)",
        "FONASA (Salud)",
        "IRPF",
        "Fondo de Solidaridad",
        "Otras deducciones",
    ]
    for i, item in enumerate(deducciones):
        r = r_ded_h + 1 + i
        write_label(ws, r, 2, item)
        style_data_cell(ws, r, 3)
        style_data_cell(ws, r, 4, currency_format)
        style_data_cell(ws, r, 5, currency_format)
        c = ws.cell(row=r, column=6)
        c.value = f"=D{r}+E{r}"
        c.number_format = currency_format
        c.border = thin_border
        c.alignment = Alignment(horizontal="center")
        alt_row_fill(ws, r, 2, 6)

    r_total_ded = r_ded_h + 1 + len(deducciones)
    ws.cell(row=r_total_ded, column=2, value="TOTAL DEDUCCIONES").font = Font(bold=True, size=11, color=ACCENT_RED)
    for col in [4, 5]:
        c = ws.cell(row=r_total_ded, column=col)
        c.value = f"=SUM({get_column_letter(col)}{r_ded_h+1}:{get_column_letter(col)}{r_total_ded-1})"
        c.number_format = currency_format
        c.font = Font(bold=True, color=ACCENT_RED)
        c.border = thin_border
        c.alignment = Alignment(horizontal="center")
    c = ws.cell(row=r_total_ded, column=6)
    c.value = f"=D{r_total_ded}+E{r_total_ded}"
    c.number_format = currency_format
    c.font = Font(bold=True, color=ACCENT_RED)
    c.border = thin_border
    c.alignment = Alignment(horizontal="center")
    style_header_row(ws, r_total_ded, 2, 6)

    # Ingreso Neto - row referenced by Dashboard (row 10 area)
    # We need this to be at a predictable row. Let's calculate.
    # row=4 header, 5 items (5-9), total bruto=10, blank=11, deduccion title=12,
    # deduccion header=13, 5 items (14-18), total ded=19
    # Neto = row 21
    r_neto = r_total_ded + 2
    ws.cell(row=r_neto, column=2, value="INGRESO NETO MENSUAL").font = Font(bold=True, size=12, color=ACCENT_GREEN)
    for col in [4, 5]:
        c = ws.cell(row=r_neto, column=col)
        c.value = f"={get_column_letter(col)}{r_total_bruto}-{get_column_letter(col)}{r_total_ded}"
        c.number_format = currency_format
        c.font = Font(bold=True, size=12, color=ACCENT_GREEN)
        c.border = thin_border
        c.alignment = Alignment(horizontal="center")
    c = ws.cell(row=r_neto, column=6)
    c.value = f"=D{r_neto}+E{r_neto}"
    c.number_format = currency_format
    c.font = Font(bold=True, size=12, color=ACCENT_GREEN)
    c.border = thin_border
    c.alignment = Alignment(horizontal="center")

    # The Dashboard references Ingresos!D10, E10, F10 for net income.
    # r_neto is actually row 21. We need to adjust Dashboard references.
    # We'll store this for later adjustment.
    ws._neto_row = r_neto

    ws.sheet_view.showGridLines = False
    return r_neto


# ═════════════════════════════════════════════════════════════════════════════
#  SHEET 3: GASTOS
# ═════════════════════════════════════════════════════════════════════════════
def create_gastos(wb):
    ws = wb.create_sheet("Gastos")
    ws.sheet_properties.tabColor = ACCENT_RED
    set_col_widths(ws, [3, 35, 5, 18, 18, 18, 5, 22])

    ws.merge_cells("B2:H2")
    cell = ws.cell(row=2, column=2, value="GASTOS MENSUALES DETALLADOS")
    cell.font = title_font
    cell.alignment = Alignment(horizontal="center")

    row = 4
    headers = ["Categoría / Concepto", "", "Ella", "Él", "Compartido", "", "Total"]
    for i, h in enumerate(headers):
        ws.cell(row=row, column=2 + i, value=h)
    style_header_row(ws, row, 2, 8)

    # Gastos categories
    categories = {
        "VIVIENDA": [
            "Alquiler",
            "Gastos comunes / Expensas",
            "Electricidad (UTE)",
            "Agua (OSE)",
            "Internet / Cable",
            "Gas",
            "Seguro hogar",
        ],
        "ALIMENTACIÓN": [
            "Supermercado / Almacén",
            "Feria / Verdulería",
            "Delivery / Comer afuera",
            "Agua / Bebidas",
        ],
        "TRANSPORTE": [
            "Combustible",
            "Estacionamiento",
            "Boleto / STM",
            "Mantenimiento vehículo",
            "Seguro vehículo",
        ],
        "SALUD": [
            "Mutualista / Emergencia médica",
            "Medicamentos",
            "Odontología",
            "Gimnasio / Deporte",
        ],
        "EDUCACIÓN Y FORMACIÓN": [
            "Cursos / Congresos",
            "Libros / Material",
            "Cuotas postgrado",
        ],
        "PERSONAL Y OCIO": [
            "Ropa",
            "Higiene / Cuidado personal",
            "Entretenimiento / Salidas",
            "Suscripciones (Netflix, Spotify, etc.)",
            "Celular",
        ],
        "FINANCIERO": [
            "Cuota préstamo / Tarjeta",
            "Seguros de vida",
            "Otros financieros",
        ],
    }

    current_row = row + 1
    cat_total_rows = []

    for cat_name, items in categories.items():
        # Category header
        c = ws.cell(row=current_row, column=2, value=cat_name)
        c.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
        cat_fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
        for col in range(2, 9):
            ws.cell(row=current_row, column=col).fill = cat_fill
            ws.cell(row=current_row, column=col).border = thin_border
        c.fill = cat_fill
        current_row += 1

        first_item_row = current_row
        for item in items:
            write_label(ws, current_row, 2, f"  {item}")
            style_data_cell(ws, current_row, 3)  # spacer
            style_data_cell(ws, current_row, 4, currency_format)
            style_data_cell(ws, current_row, 5, currency_format)
            style_data_cell(ws, current_row, 6, currency_format)
            style_data_cell(ws, current_row, 7)  # spacer
            c = ws.cell(row=current_row, column=8)
            c.value = f"=D{current_row}+E{current_row}+F{current_row}"
            c.number_format = currency_format
            c.border = thin_border
            c.alignment = Alignment(horizontal="center")
            alt_row_fill(ws, current_row, 2, 8)
            current_row += 1

        # Subtotal row
        last_item_row = current_row - 1
        ws.cell(row=current_row, column=2, value=f"  Subtotal {cat_name.title()}").font = Font(bold=True, size=10, italic=True)
        for col in [4, 5, 6, 8]:
            letter = get_column_letter(col)
            c = ws.cell(row=current_row, column=col)
            c.value = f"=SUM({letter}{first_item_row}:{letter}{last_item_row})"
            c.number_format = currency_format
            c.font = Font(bold=True, size=10, italic=True)
            c.border = thin_border
            c.alignment = Alignment(horizontal="center")
        sub_fill = PatternFill(start_color=SOFT_BLUE, end_color=SOFT_BLUE, fill_type="solid")
        for col in range(2, 9):
            ws.cell(row=current_row, column=col).fill = sub_fill
            ws.cell(row=current_row, column=col).border = thin_border
        cat_total_rows.append(current_row)
        current_row += 1

    # GRAN TOTAL
    current_row += 1
    total_row = current_row
    ws.cell(row=total_row, column=2, value="TOTAL GASTOS MENSUALES").font = Font(bold=True, size=12, color=DARK_NAVY)
    for col in [4, 5, 6, 8]:
        letter = get_column_letter(col)
        sum_refs = "+".join([f"{letter}{r}" for r in cat_total_rows])
        c = ws.cell(row=total_row, column=col)
        c.value = f"={sum_refs}"
        c.number_format = currency_format
        c.font = Font(bold=True, size=12)
        c.border = thin_border
        c.alignment = Alignment(horizontal="center")
    style_header_row(ws, total_row, 2, 8)

    ws._total_row = total_row
    ws.sheet_view.showGridLines = False
    return total_row


# ═════════════════════════════════════════════════════════════════════════════
#  SHEET 4: OBJETIVOS DE AHORRO
# ═════════════════════════════════════════════════════════════════════════════
def create_objetivos(wb):
    ws = wb.create_sheet("Objetivos")
    ws.sheet_properties.tabColor = ACCENT_YELLOW
    set_col_widths(ws, [3, 30, 20, 20, 20, 20, 18, 20, 20])

    ws.merge_cells("B2:I2")
    cell = ws.cell(row=2, column=2, value="OBJETIVOS DE AHORRO E INVERSIÓN")
    cell.font = title_font
    cell.alignment = Alignment(horizontal="center")

    ws.merge_cells("B3:I3")
    cell = ws.cell(row=3, column=2, value="Definir metas claras con plazos y montos específicos")
    cell.font = Font(name="Calibri", size=10, italic=True, color=DARK_GRAY)
    cell.alignment = Alignment(horizontal="center")

    row = 5
    headers = [
        "Objetivo", "Plazo", "Meta Total ($)",
        "Ahorrado ($)", "Restante ($)", "Ahorro Mensual Req.",
        "Fecha Inicio", "Fecha Meta"
    ]
    for i, h in enumerate(headers):
        ws.cell(row=row, column=2 + i, value=h)
    style_header_row(ws, row, 2, 9)

    objectives = [
        ("1. Vacaciones Juntos", "Corto (6-12 meses)"),
        ("2. Ahorro Individual - Ella", "Corto-Largo"),
        ("3. Ahorro Individual - Él", "Corto-Largo"),
        ("4. Compra de Inmueble", "Mediano (3-5 años)"),
        ("5. Retiro / Inversiones", "Largo (10+ años)"),
    ]

    for i, (obj, plazo) in enumerate(objectives):
        r = row + 1 + i
        write_label(ws, r, 2, obj, bold=True)
        write_label(ws, r, 3, plazo)
        # Meta (user fills)
        style_data_cell(ws, r, 4, currency_format)
        # Ahorrado (user fills or linked to Seguimiento)
        style_data_cell(ws, r, 5, currency_format)
        # Restante
        c = ws.cell(row=r, column=6)
        c.value = f"=D{r}-E{r}"
        c.number_format = currency_format
        c.border = thin_border
        c.alignment = Alignment(horizontal="center")
        # Ahorro mensual requerido (placeholder formula)
        c = ws.cell(row=r, column=7)
        c.value = f'=IF(AND(H{r}<>"",I{r}<>""),F{r}/MAX((I{r}-H{r})/30,1),0)'
        c.number_format = currency_format
        c.border = thin_border
        c.alignment = Alignment(horizontal="center")
        # Fecha inicio
        style_data_cell(ws, r, 8, "DD/MM/YYYY")
        # Fecha meta
        style_data_cell(ws, r, 9, "DD/MM/YYYY")
        alt_row_fill(ws, r, 2, 9)

    # ── Detalle por objetivo ──
    row_detail = row + 1 + len(objectives) + 2

    # Objective 1: Vacaciones
    write_section_title(ws, row_detail, 2, "Objetivo 1: Vacaciones Juntos (Corto Plazo)")
    row_detail += 1
    detail_labels_vac = [
        ("Destino", ""),
        ("Duración (días)", ""),
        ("Presupuesto vuelos/transporte", ""),
        ("Presupuesto alojamiento", ""),
        ("Presupuesto comida/actividades", ""),
        ("Presupuesto extras/imprevistos", ""),
        ("TOTAL ESTIMADO", "=SUM(C{start}:C{end})"),
    ]
    vac_start_row = row_detail + 2
    for i, (label, _) in enumerate(detail_labels_vac):
        r = row_detail + i
        write_label(ws, r, 2, label, bold=(i == len(detail_labels_vac) - 1))
        c = style_data_cell(ws, r, 3, currency_format)
        if i == 0 or i == 1:
            c.number_format = "General"
        if i == len(detail_labels_vac) - 1:
            c.value = f"=SUM(C{vac_start_row}:C{r-1})"
            c.font = Font(bold=True)

    # Objective 2 & 3: Ahorro Individual
    row_detail += len(detail_labels_vac) + 2
    write_section_title(ws, row_detail, 2, "Objetivos 2-3: Ahorro Individual")
    row_detail += 1
    indiv_items = [
        "Fondo de emergencia personal",
        "Desarrollo profesional",
        "Gustos personales",
        "Otros",
    ]
    headers_indiv = ["Concepto", "Ella ($)", "Él ($)"]
    for i, h in enumerate(headers_indiv):
        ws.cell(row=row_detail, column=2 + i, value=h)
    style_header_row(ws, row_detail, 2, 4)
    row_detail += 1
    for i, item in enumerate(indiv_items):
        r = row_detail + i
        write_label(ws, r, 2, item)
        style_data_cell(ws, r, 3, currency_format)
        style_data_cell(ws, r, 4, currency_format)
        alt_row_fill(ws, r, 2, 4)

    # Objective 4: Inmueble
    row_detail += len(indiv_items) + 2
    write_section_title(ws, row_detail, 2, "Objetivo 4: Compra de Inmueble (Mediano Plazo)")
    row_detail += 1
    inmueble_items = [
        ("Valor estimado del inmueble", ""),
        ("% de entrada requerido", ""),
        ("Monto de entrada necesario", ""),
        ("Gastos escrituración/notarial", ""),
        ("Refacción / Equipamiento", ""),
        ("TOTAL A AHORRAR", ""),
    ]
    for i, (label, _) in enumerate(inmueble_items):
        r = row_detail + i
        write_label(ws, r, 2, label, bold=(i == len(inmueble_items) - 1))
        c = style_data_cell(ws, r, 3, currency_format)
        if i == 1:
            c.number_format = pct_format

    # Objective 5: Retiro
    row_detail += len(inmueble_items) + 2
    write_section_title(ws, row_detail, 2, "Objetivo 5: Retiro e Inversiones (Largo Plazo)")
    row_detail += 1
    retiro_items = [
        "AFAP voluntario",
        "Fondo de inversión (UI / USD)",
        "Letras de regulación monetaria",
        "Bonos del tesoro",
        "Inversiones en acciones / ETFs",
        "Otros instrumentos",
    ]
    headers_ret = ["Instrumento", "Aporte Mensual ($)", "Rentabilidad Est. Anual"]
    for i, h in enumerate(headers_ret):
        ws.cell(row=row_detail, column=2 + i, value=h)
    style_header_row(ws, row_detail, 2, 4)
    row_detail += 1
    for i, item in enumerate(retiro_items):
        r = row_detail + i
        write_label(ws, r, 2, item)
        style_data_cell(ws, r, 3, currency_format)
        style_data_cell(ws, r, 4, pct_format)
        alt_row_fill(ws, r, 2, 4)

    ws.sheet_view.showGridLines = False


# ═════════════════════════════════════════════════════════════════════════════
#  SHEET 5: SEGUIMIENTO MENSUAL
# ═════════════════════════════════════════════════════════════════════════════
def create_seguimiento(wb):
    ws = wb.create_sheet("Seguimiento")
    ws.sheet_properties.tabColor = LIGHT_BLUE
    set_col_widths(ws, [3, 18] + [16] * 12 + [18])

    ws.merge_cells("B2:N2")
    cell = ws.cell(row=2, column=2, value="SEGUIMIENTO MENSUAL - AÑO 2025")
    cell.font = title_font
    cell.alignment = Alignment(horizontal="center")

    row = 4
    headers = ["Concepto"] + MESES + ["TOTAL ANUAL"]
    for i, h in enumerate(headers):
        ws.cell(row=row, column=2 + i, value=h)
    style_header_row(ws, row, 2, 15)

    sections = [
        ("INGRESOS", [
            "Ingreso Neto Ella",
            "Ingreso Neto Él",
        ], ACCENT_GREEN),
        ("GASTOS", [
            "Vivienda",
            "Alimentación",
            "Transporte",
            "Salud",
            "Educación",
            "Personal/Ocio",
            "Financiero",
        ], ACCENT_RED),
        ("AHORRO POR OBJETIVO", [
            "Vacaciones",
            "Ahorro Individual Ella",
            "Ahorro Individual Él",
            "Inmueble",
            "Retiro / Inversiones",
        ], ACCENT_YELLOW),
    ]

    current_row = row + 1
    section_total_refs = {}

    for sec_name, items, color in sections:
        # Section header
        c = ws.cell(row=current_row, column=2, value=sec_name)
        c.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
        sec_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for col in range(2, 16):
            ws.cell(row=current_row, column=col).fill = sec_fill
            ws.cell(row=current_row, column=col).border = thin_border
        current_row += 1

        first_item_row = current_row
        for item in items:
            write_label(ws, current_row, 2, f"  {item}")
            for col in range(3, 15):
                style_data_cell(ws, current_row, col, currency_format)
            # Total anual
            c = ws.cell(row=current_row, column=15)
            c.value = f"=SUM(C{current_row}:N{current_row})"
            c.number_format = currency_format
            c.border = thin_border
            c.alignment = Alignment(horizontal="center")
            c.font = Font(bold=True)
            alt_row_fill(ws, current_row, 2, 15)
            current_row += 1

        # Section total
        last_item_row = current_row - 1
        ws.cell(row=current_row, column=2, value=f"TOTAL {sec_name}").font = Font(bold=True, size=10)
        for col in range(3, 16):
            letter = get_column_letter(col)
            c = ws.cell(row=current_row, column=col)
            c.value = f"=SUM({letter}{first_item_row}:{letter}{last_item_row})"
            c.number_format = currency_format
            c.font = Font(bold=True)
            c.border = thin_border
            c.alignment = Alignment(horizontal="center")
        sub_fill = PatternFill(start_color=SOFT_BLUE, end_color=SOFT_BLUE, fill_type="solid")
        for col in range(2, 16):
            ws.cell(row=current_row, column=col).fill = sub_fill
            ws.cell(row=current_row, column=col).border = thin_border
        section_total_refs[sec_name] = current_row
        current_row += 1

    # Balance row
    current_row += 1
    ws.cell(row=current_row, column=2, value="BALANCE MENSUAL").font = Font(bold=True, size=11, color=DARK_NAVY)
    ing_row = section_total_refs["INGRESOS"]
    gas_row = section_total_refs["GASTOS"]
    aho_row = section_total_refs["AHORRO POR OBJETIVO"]
    for col in range(3, 16):
        letter = get_column_letter(col)
        c = ws.cell(row=current_row, column=col)
        c.value = f"={letter}{ing_row}-{letter}{gas_row}-{letter}{aho_row}"
        c.number_format = currency_format
        c.font = Font(bold=True, size=11)
        c.border = thin_border
        c.alignment = Alignment(horizontal="center")
    style_header_row(ws, current_row, 2, 15)

    # Tasa de ahorro row
    current_row += 1
    ws.cell(row=current_row, column=2, value="TASA DE AHORRO").font = Font(bold=True, size=10, color=ACCENT_GREEN)
    for col in range(3, 16):
        letter = get_column_letter(col)
        c = ws.cell(row=current_row, column=col)
        c.value = f'=IF({letter}{ing_row}>0,{letter}{aho_row}/{letter}{ing_row},0)'
        c.number_format = pct_format
        c.font = Font(bold=True, color=ACCENT_GREEN)
        c.border = thin_border
        c.alignment = Alignment(horizontal="center")

    # Acumulado ahorro
    current_row += 2
    ws.cell(row=current_row, column=2, value="AHORRO ACUMULADO").font = Font(bold=True, size=11, color=DARK_NAVY)
    for col in range(3, 16):
        letter = get_column_letter(col)
        c = ws.cell(row=current_row, column=col)
        if col == 3:
            c.value = f"={letter}{aho_row}"
        else:
            prev_letter = get_column_letter(col - 1)
            c.value = f"={prev_letter}{current_row}+{letter}{aho_row}"
        c.number_format = currency_format
        c.font = Font(bold=True)
        c.border = thin_border
        c.alignment = Alignment(horizontal="center")

    ws.sheet_view.showGridLines = False


# ═════════════════════════════════════════════════════════════════════════════
#  SHEET 6: INVERSIONES
# ═════════════════════════════════════════════════════════════════════════════
def create_inversiones(wb):
    ws = wb.create_sheet("Inversiones")
    ws.sheet_properties.tabColor = ACCENT_ORANGE
    set_col_widths(ws, [3, 28, 18, 18, 16, 18, 18, 22])

    ws.merge_cells("B2:H2")
    cell = ws.cell(row=2, column=2, value="PORTAFOLIO DE INVERSIONES")
    cell.font = title_font
    cell.alignment = Alignment(horizontal="center")

    ws.merge_cells("B3:H3")
    cell = ws.cell(row=3, column=2, value="Opciones disponibles en Uruguay para residentes médicos")
    cell.font = Font(name="Calibri", size=10, italic=True, color=DARK_GRAY)
    cell.alignment = Alignment(horizontal="center")

    # ── Opciones de inversión en Uruguay ──
    row = 5
    write_section_title(ws, row, 2, "Opciones de Inversión - Uruguay")

    row = 6
    headers = ["Instrumento", "Riesgo", "Liquidez", "Moneda", "Rent. Est. Anual", "Aporte Mensual", "Notas"]
    for i, h in enumerate(headers):
        ws.cell(row=row, column=2 + i, value=h)
    style_header_row(ws, row, 2, 8)

    instruments = [
        ("AFAP Voluntaria (República/Unión)", "Bajo", "Baja", "UI/UYU", "3-5%", "", "Deducible de IRPF"),
        ("Letras de Regulación Monetaria", "Muy bajo", "Media", "UYU", "Variable BCU", "", "Plazo 30-360 días"),
        ("Notas del Tesoro en UI", "Bajo", "Media", "UI", "2-4% real", "", "Protege contra inflación"),
        ("Bonos Globales Uruguay", "Bajo", "Media", "USD", "4-6%", "", "Renta fija en USD"),
        ("Depósito a plazo fijo", "Muy bajo", "Baja", "UYU/USD", "Variable", "", "Garantía BCU"),
        ("Fondo de inversión (SURA, etc.)", "Medio", "Media", "USD/UI", "5-8%", "", "Diversificado"),
        ("Acciones / ETFs (broker int'l)", "Alto", "Alta", "USD", "8-12%", "", "Requiere cuenta exterior"),
        ("Inmueble (renta)", "Medio", "Muy baja", "UYU/USD", "4-7%", "", "Ingreso pasivo"),
    ]

    for i, (inst, risk, liq, mon, rent, aporte, notas) in enumerate(instruments):
        r = row + 1 + i
        write_label(ws, r, 2, inst)
        write_label(ws, r, 3, risk)
        write_label(ws, r, 4, liq)
        write_label(ws, r, 5, mon)
        write_label(ws, r, 6, rent)
        style_data_cell(ws, r, 7, currency_format)
        write_label(ws, r, 8, notas)
        # Color risk
        risk_cell = ws.cell(row=r, column=3)
        if "Muy bajo" in risk:
            risk_cell.font = Font(color=ACCENT_GREEN, size=10)
        elif "Bajo" == risk:
            risk_cell.font = Font(color="2E7D32", size=10)
        elif "Medio" in risk:
            risk_cell.font = Font(color=ACCENT_ORANGE, size=10)
        elif "Alto" in risk:
            risk_cell.font = Font(color=ACCENT_RED, size=10)
        alt_row_fill(ws, r, 2, 8)

    # Total
    r_total = row + 1 + len(instruments)
    ws.cell(row=r_total, column=2, value="TOTAL INVERSIÓN MENSUAL").font = Font(bold=True, size=11)
    c = ws.cell(row=r_total, column=7)
    c.value = f"=SUM(G{row+1}:G{r_total-1})"
    c.number_format = currency_format
    c.font = Font(bold=True)
    c.border = thin_border
    c.alignment = Alignment(horizontal="center")
    style_header_row(ws, r_total, 2, 8)

    # ── Simulador de interés compuesto ──
    row_sim = r_total + 3
    write_section_title(ws, row_sim, 2, "Simulador de Interés Compuesto")

    row_sim += 1
    sim_inputs = [
        ("Aporte mensual ($)", currency_format),
        ("Tasa anual esperada (%)", pct_format),
        ("Años de inversión", "General"),
    ]
    for i, (label, fmt) in enumerate(sim_inputs):
        r = row_sim + i
        write_label(ws, r, 2, label, bold=True)
        style_data_cell(ws, r, 3, fmt)

    aporte_cell = f"C{row_sim}"
    tasa_cell = f"C{row_sim+1}"
    anos_cell = f"C{row_sim+2}"

    row_sim += len(sim_inputs) + 1
    write_label(ws, row_sim, 2, "Total Invertido (sin rendimiento)", bold=True)
    c = ws.cell(row=row_sim, column=3)
    c.value = f"={aporte_cell}*{anos_cell}*12"
    c.number_format = currency_format
    c.font = Font(bold=True)
    c.border = thin_border
    c.alignment = Alignment(horizontal="center")

    row_sim += 1
    write_label(ws, row_sim, 2, "Valor Futuro (con rendimiento)", bold=True)
    c = ws.cell(row=row_sim, column=3)
    # FV of annuity formula: PMT * (((1+r/12)^(n*12) - 1) / (r/12))
    c.value = f'=IF(AND({aporte_cell}>0,{tasa_cell}>0,{anos_cell}>0),{aporte_cell}*(((1+{tasa_cell}/12)^({anos_cell}*12)-1)/({tasa_cell}/12)),0)'
    c.number_format = currency_format
    c.font = Font(bold=True, size=12, color=ACCENT_GREEN)
    c.border = thin_border
    c.alignment = Alignment(horizontal="center")

    row_sim += 1
    write_label(ws, row_sim, 2, "Ganancia por Rendimiento", bold=True)
    c = ws.cell(row=row_sim, column=3)
    c.value = f"=C{row_sim-1}-C{row_sim-2}"
    c.number_format = currency_format
    c.font = Font(bold=True, color=ACCENT_GREEN)
    c.border = thin_border
    c.alignment = Alignment(horizontal="center")

    # ── Perfil de riesgo sugerido ──
    row_perf = row_sim + 3
    write_section_title(ws, row_perf, 2, "Perfil de Riesgo Sugerido para Médicos Residentes")
    row_perf += 1
    tips = [
        "• Prioridad 1: Fondo de emergencia (3-6 meses gastos) en depósito o LRM.",
        "• Prioridad 2: AFAP voluntaria para beneficio fiscal (deducción IRPF).",
        "• Prioridad 3: Notas en UI para ahorro mediano plazo (inmueble).",
        "• Prioridad 4: Diversificar en fondos USD para largo plazo.",
        "• Cuidado: Evitar inversiones ilíquidas mientras están en residencia.",
        "• Tip: Automatizar transferencias el día de cobro.",
    ]
    for i, tip in enumerate(tips):
        c = ws.cell(row=row_perf + i, column=2, value=tip)
        c.font = Font(name="Calibri", size=10, color=DARK_GRAY)

    ws.sheet_view.showGridLines = False


# ═════════════════════════════════════════════════════════════════════════════
#  SHEET 7: FONDO DE EMERGENCIA
# ═════════════════════════════════════════════════════════════════════════════
def create_fondo_emergencia(wb):
    ws = wb.create_sheet("Fondo Emergencia")
    ws.sheet_properties.tabColor = "8B0000"
    set_col_widths(ws, [3, 35, 22, 22])

    ws.merge_cells("B2:D2")
    cell = ws.cell(row=2, column=2, value="FONDO DE EMERGENCIA")
    cell.font = title_font
    cell.alignment = Alignment(horizontal="center")

    ws.merge_cells("B3:D3")
    cell = ws.cell(row=3, column=2, value="Meta: cubrir 3 a 6 meses de gastos esenciales")
    cell.font = Font(name="Calibri", size=10, italic=True, color=DARK_GRAY)
    cell.alignment = Alignment(horizontal="center")

    row = 5
    write_section_title(ws, row, 2, "Cálculo de Meta")
    row += 1

    items = [
        ("Gastos esenciales mensuales", "=Gastos!H43", currency_format),
        ("Meta 3 meses", "=C6*3", currency_format),
        ("Meta 6 meses (recomendado)", "=C6*6", currency_format),
        ("Ahorrado actualmente", "", currency_format),
        ("Faltan para meta 3 meses", "=MAX(C7-C9,0)", currency_format),
        ("Faltan para meta 6 meses", "=MAX(C8-C9,0)", currency_format),
        ("Ahorro mensual para este fondo", "", currency_format),
        ("Meses para alcanzar meta 3m", '=IF(C12>0,C10/C12,"—")', "General"),
        ("Meses para alcanzar meta 6m", '=IF(C12>0,C11/C12,"—")', "General"),
    ]

    for i, (label, formula, fmt) in enumerate(items):
        r = row + i
        write_label(ws, r, 2, label, bold=(i in [1, 2, 5]))
        c = ws.cell(row=r, column=3)
        if formula:
            c.value = formula
        c.number_format = fmt
        c.border = thin_border
        c.alignment = Alignment(horizontal="center")
        alt_row_fill(ws, r, 2, 3)

    # Where to keep it
    row_where = row + len(items) + 2
    write_section_title(ws, row_where, 2, "¿Dónde guardar el fondo de emergencia?")
    row_where += 1
    options = [
        "• Cuenta de ahorro en UYU (acceso inmediato)",
        "• Depósito a plazo 30 días (mejor tasa, algo menos líquido)",
        "• Letras de regulación monetaria corto plazo",
        "• NO en inversiones de riesgo o poco líquidas",
        "",
        "IMPORTANTE: Este fondo es intocable salvo emergencia real.",
        "No es para vacaciones, no es para oportunidades de inversión.",
    ]
    for i, opt in enumerate(options):
        c = ws.cell(row=row_where + i, column=2, value=opt)
        c.font = Font(name="Calibri", size=10, color=DARK_GRAY if i < 4 else ACCENT_RED)

    ws.sheet_view.showGridLines = False


# ═════════════════════════════════════════════════════════════════════════════
#  MAIN: Build workbook
# ═════════════════════════════════════════════════════════════════════════════
def main():
    wb = openpyxl.Workbook()

    # Create all sheets
    create_dashboard(wb)
    neto_row = create_ingresos(wb)
    total_gastos_row = create_gastos(wb)
    create_objetivos(wb)
    create_seguimiento(wb)
    create_inversiones(wb)
    create_fondo_emergencia(wb)

    # ── Fix Dashboard references to match actual rows ──
    ws_dash = wb["Dashboard"]

    # Ingresos references - update to actual neto row
    ws_dash.cell(row=6, column=3).value = f"=Ingresos!D{neto_row}"
    ws_dash.cell(row=7, column=3).value = f"=Ingresos!E{neto_row}"
    ws_dash.cell(row=8, column=3).value = f"=Ingresos!F{neto_row}"

    # Gastos reference
    ws_dash.cell(row=9, column=3).value = f"=Gastos!H{total_gastos_row}"

    # Freeze panes on key sheets
    wb["Seguimiento"].freeze_panes = "C5"
    wb["Gastos"].freeze_panes = "C5"

    # Print settings
    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        ws.page_setup.orientation = "landscape"
        ws.page_setup.paperSize = ws.PAPERSIZE_A4

    output_path = "/home/user/mckako/Plan_Financiero_Residentes_Uruguay.xlsx"
    wb.save(output_path)
    print(f"Archivo generado exitosamente: {output_path}")


if __name__ == "__main__":
    main()
